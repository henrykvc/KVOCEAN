"""
Convert OCR 매핑 검토용 시트.xlsx -> src/lib/validation/classification-seed.ts

Excel layout (1-indexed columns):
  A: ㅇ (sequence)         B: 비고2                 C: 대분류
  D: 일치값 (=중분류)       E: 중분류 (=소분류)       F: 양/음 텍스트(헤더 마커)
  G: 계정명 (=세분류)       H: -                     I: 소분류 (=항목/alias)
  J: 넘버 (코드)            K: 양/음 코드 (0/1)

Output: TypeScript module with seeded classification entries.
One entry per (code, sign) combination since 음수 is a separate code suffix.
"""
import json
import sys
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r"C:\Users\aagyw\Downloads\_1_ OCR매핑_검토용 시트 (3).xlsx")
OUT = Path(__file__).parent.parent / "src" / "lib" / "validation" / "classification-seed.ts"


def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Mapping DB_COA"]

    # code -> {meta: {대, 중, 소, 세}, sign, aliases: [str]}
    entries = {}

    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 10).value
        if code is None:
            continue
        대 = ws.cell(r, 3).value
        중 = ws.cell(r, 4).value
        소 = ws.cell(r, 5).value
        세 = ws.cell(r, 7).value
        항목 = ws.cell(r, 9).value
        sign = ws.cell(r, 11).value
        sign = 1 if sign == 1 else 0

        if code not in entries:
            entries[code] = {
                "code": code,
                "대": (대 or "").strip() or None,
                "중": (중 or "").strip() or None,
                "소": (소 or "").strip() or None,
                "세": (세 or "").strip() or None,
                "sign": sign,
                "aliases": [],
            }

        e = entries[code]
        if 대 and not e["대"]: e["대"] = 대.strip()
        if 중 and not e["중"]: e["중"] = 중.strip()
        if 소 and not e["소"]: e["소"] = 소.strip()
        if 세 and not e["세"]: e["세"] = 세.strip()

        # Collect aliases: G(세분류 자체) and I(소분류=항목)
        if 세:
            v = 세.strip()
            if v and v not in e["aliases"]:
                e["aliases"].append(v)
        if 항목:
            v = 항목.strip()
            if v and v not in e["aliases"]:
                e["aliases"].append(v)

    # Normalize: backfill missing 대/중/소 from the +/− paired code.
    # 음수 페어(코드 +1000)에서 메타가 비어있으면 양수 페어에서 빌려옴.
    for code, e in entries.items():
        if e["대"] and e["중"] and e["소"]:
            continue
        # Try paired code: same prefix, opposite sign suffix
        for delta in (-1000, 1000):
            paired = entries.get(code + delta)
            if not paired:
                continue
            if not e["대"] and paired["대"]: e["대"] = paired["대"]
            if not e["중"] and paired["중"]: e["중"] = paired["중"]
            if not e["소"] and paired["소"]: e["소"] = paired["소"]
            if not e["세"] and paired["세"]: e["세"] = paired["세"]

    cleaned = []
    skipped = []
    for code in sorted(entries):
        e = entries[code]
        if not e["대"]:
            skipped.append(e)
            continue
        if not e["세"]:
            e["세"] = (e["aliases"][0] if e["aliases"] else e["소"]) or ""
        cleaned.append(e)

    # Manual augmentation: legacy ACCOUNT_ALIASES variants that weren't in the
    # source xlsx, attached to verified seed codes so single-source matching
    # (no keyword fallback) still recognizes common label variants.
    #
    # Aggregate names (e.g. "판매비와관리비 합계", "영업외수익 합계") are NOT
    # attached here — they belong to SUMMARY_RULES code-prefix aggregation,
    # not to any single 세분류 entry.
    EXTRA_ALIASES = {
        3048000: ["결손금"],
        4011000: ["영업이익(손익)", "영업이익또는손실", "영업이익(영업손실)", "영업이익(손실)"],
        4068000: ["세전계속사업이익", "법인세차감전이익(손실)", "법인세비용차감전순이익(손실)", "법인세비용차감전계속사업이익"],
        4070000: ["법인세 등", "법인세수익", "당기법인세비용", "이연법인세비용", "법인세환급"],
        4071000: ["당기순이익(손실)", "당기순손익", "연결당기순이익", "당기순이익(당기순손실)"],
    }

    for code, extras in EXTRA_ALIASES.items():
        match = next((c for c in cleaned if c["code"] == code), None)
        if not match:
            print(f'  ! EXTRA_ALIASES: code {code} not in cleaned — skipping {extras}')
            continue
        for a in extras:
            if a not in match["aliases"]:
                match["aliases"].append(a)

    print(f"Total codes:   {len(entries)}")
    print(f"Kept entries:  {len(cleaned)}")
    print(f"Skipped (no 대분류): {len(skipped)}")
    if skipped:
        print("  e.g.", skipped[:3])

    # Render TS module
    out = []
    out.append("// AUTO-GENERATED from _1_ OCR매핑_검토용 시트 (3).xlsx")
    out.append("// Source: scripts/convert-classification-seed.py")
    out.append("// DO NOT EDIT BY HAND.")
    out.append("")
    out.append("export type ClassificationSeedEntry = {")
    out.append("  code: number;        // 7-digit composite key")
    out.append("  대분류: string;")
    out.append("  중분류: string;       // OCR section name (file's 일치값)")
    out.append("  소분류: string;       // sub-group (file's 중분류)")
    out.append("  세분류: string;       // canonical leaf name (file's 계정명)")
    out.append("  sign: 0 | 1;         // 0 = +, 1 = − (already baked into code suffix)")
    out.append("  aliases: string[];   // OCR raw account name variants")
    out.append("};")
    out.append("")
    out.append("export const CLASSIFICATION_SEED: ClassificationSeedEntry[] = [")
    for e in cleaned:
        대 = json.dumps(e["대"], ensure_ascii=False)
        중 = json.dumps(e["중"] or "", ensure_ascii=False)
        소 = json.dumps(e["소"] or "", ensure_ascii=False)
        세 = json.dumps(e["세"] or "", ensure_ascii=False)
        aliases = json.dumps(e["aliases"], ensure_ascii=False)
        out.append(
            f"  {{ code: {e['code']}, 대분류: {대}, 중분류: {중}, 소분류: {소}, 세분류: {세}, sign: {e['sign']}, aliases: {aliases} }},"
        )
    out.append("];")
    out.append("")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(out), encoding="utf-8")
    print(f"Wrote: {OUT}")


if __name__ == "__main__":
    main()
